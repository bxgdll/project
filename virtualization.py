import sys
import itertools
import numpy as np
import pandas as pd
import openpyxl
import random


def main(num):
    # 执行体池
    _pool = pd.read_excel('virtualization_dispatch.xlsx', usecols=['序号', '虚拟化方式', '漏洞（1、4、6）', '漏洞（2、5、7）', '漏洞（3）', '可信度', '被调度次数', '下线次数', '指标'])
    pool = np.array(_pool)
    print("初始执行体池:")
    print('序号', '虚拟化方式', '漏洞（1、4、6）', '漏洞（2、5、7）', '漏洞（3）', '可信度', '被调度次数', '下线次数', '指标')
    print(pool)
    # 按指标排序
    pool_sort = np.array(sorted(pool, key=lambda x: x[8], reverse=True))
    print("初始化选取的", num, "个执行体:")
    print('序号', '虚拟化方式', '漏洞（1、4、6）', '漏洞（2、5、7）', '漏洞（3）', '可信度', '被调度次数', '下线次数', '指标')
    # 选择执行体
    E = pool_sort[0:num]
    print(E)
    for i in range(0, num):
        for j in range(0, 30):
            if pool[j][0] == E[i][0]:
                # 选中次数+1，然后更新可信度和指标
                pool[j][6] += 1
                pool[j][5] = ((pool[j][5] * pool[j][6]) + 1.0) / (pool[j][6] + 1.0)
                # 更新指标
                if pool[j][0] >= 0 and pool[j][0] <= 9:
                    pool[j][8] = 0.8 * pool[j][5] + 0.2 * 1.0 / 1.925926
                elif pool[j][0] >= 10 and pool[j][0] <= 19:
                    pool[j][8] = 0.8 * pool[j][5] + 0.2 * 1.0 / 2.611111
                else:
                    pool[j][8] = 0.8 * pool[j][5] + 0.2 * 1.0 / 2.5
    # 更新
    print("选取这", num, "个执行体后更新的执行体池:")
    print('序号', '虚拟化方式', '漏洞（1、4、6）', '漏洞（2、5、7）', '漏洞（3）', '可信度', '被调度次数', '下线次数', '指标')
    print(pool)
    # 存储被选择的执行体包含的漏洞
    # _holes = []
    # for i in range(0, num):
    #     _holes.append(E[i][2])
    #     _holes.append(E[i][3])
    #     _holes.append(E[i][4])
    # holes = list(set(_holes))
    # 模拟攻击
    numbers = range(1, 8)
    # 存储由于当前攻击而引发虚拟化逃逸的漏洞
    holes = []
    index = 2

    if num == 3:
        wb = openpyxl.load_workbook('result_3.xlsx')
        ws = wb['Sheet1']
        ws.cell(row=1, column=1, value="调度次数")
        ws.cell(row=1, column=2, value="指标")
    elif num == 4:
        wb = openpyxl.load_workbook('result_4.xlsx')
        ws = wb['Sheet1']
        ws.cell(row=1, column=1, value="调度次数")
        ws.cell(row=1, column=2, value="指标")
    else:
        wb = openpyxl.load_workbook('result_5.xlsx')
        ws = wb['Sheet1']
        ws.cell(row=1, column=1, value="调度次数")
        ws.cell(row=1, column=2, value="指标")

    for i in range(1, 8):
        attacks = itertools.combinations(numbers, i)
        # attack是元组
        for attack in attacks:
            for k in range(0, num):
                for j in attack:
                    for m in range(2, 5):
                        if E[k][m] == j:
                            # 存储和当前攻击相关的漏洞
                            holes.append(E[k][m])
                # 发生虚拟化逃逸
                if holes != []:
                    # 清洗+调度
                    for n in range(0, 30):
                        # 清洗下线
                        if pool[n][0] == E[k][0]:
                            pool[n][7] += 1
                    # 选取不包含当前漏洞且指标值最大的那个
                    diaodu = []
                    judge = 1
                    for n in range(0, 30):
                        for h in holes:
                            if pool[n][2] == h or pool[n][3] == h or pool[n][4] == h:
                                # 不作为调度的评选对象
                                judge = 0
                                break
                        if judge == 1:
                            diaodu.append(pool[n])
                        judge = 1
                    # 按指标排序
                    if diaodu != []:
                        diaodu_sort = np.array(sorted(diaodu, key=lambda x: x[8], reverse=True))
                        E[k] = diaodu_sort[0]
                        ws.cell(row=index, column=1).value = index - 1
                        ws.cell(row=index, column=2).value = diaodu_sort[0][8]
                        index += 1
                        if num == 3:
                            wb.save('result_3.xlsx')
                        elif num == 4:
                            wb.save('result_4.xlsx')
                        else:
                            wb.save('result_5.xlsx')

                        # 被选中，更新
                        for n in range(0, 30):
                            if pool[n][0] == E[k][0]:
                                # 选中次数+1，然后更新可信度和指标
                                pool[n][6] += 1
                                pool[n][5] = ((pool[n][5] * pool[n][6]) + 1.0) / (pool[n][6] + 1.0)
                                # 更新指标
                                if pool[n][0] >= 0 and pool[n][0] <= 9:
                                    pool[n][8] = 0.8 * pool[n][5] + 0.2 * 1.0 / 1.925926
                                elif pool[n][0] >= 10 and pool[n][0] <= 19:
                                    pool[n][8] = 0.8 * pool[n][5] + 0.2 * 1.0 / 2.611111
                                else:
                                    pool[n][8] = 0.8 * pool[n][5] + 0.2 * 1.0 / 2.5
                holes = []


    print("攻击结束后，执行体池:")
    print('序号', '虚拟化方式', '漏洞（1、4、6）', '漏洞（2、5、7）', '漏洞（3）', '可信度', '被调度次数', '下线次数', '指标')
    print(pool)

    # 存储平均指标
    wb = openpyxl.load_workbook('result.xlsx')
    ws = wb['Sheet1']
    ws.cell(row=1, column=(num - 1) * 5 - 4, value="序号")
    ws.cell(row=1, column=(num - 1) * 5 - 3, value="被调度次数")
    ws.cell(row=1, column=(num - 1) * 5 - 2, value="下线次数")
    ws.cell(row=1, column=(num - 1) * 5 - 1, value="指标")

    pool_list = pool.tolist()

    for row_index, row in enumerate(pool_list):
        ws.cell(row=row_index + 2, column=(num - 1) * 5 - 4, value=row[0])
        ws.cell(row=row_index + 2, column=(num-1)*5-3, value=row[6])
        ws.cell(row=row_index + 2, column=(num-1)*5-2, value=row[7])
        ws.cell(row=row_index + 2, column=(num-1)*5-1, value=row[8])

    wb.save('result.xlsx')



if __name__ == "__main__":
      main(3)
